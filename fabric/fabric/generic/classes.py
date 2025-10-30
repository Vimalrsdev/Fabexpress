"""
------------------------
Generic classes module
The module consists utility classes that can be used by various modules in various scenarios.
------------------------
Coded by: Krishna Prasad K
© Jyothy Fabricare Services LTD.
------------------------
"""
import os
from datetime import datetime, time, date
from decimal import Decimal

from flask import current_app, send_from_directory, json

from fabric import db

# import inflection
from .functions import calculate_distance
from .loggers import error_logger

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Alignment, Side


class GetDict:
    def as_dict(self):
        """
        Generating a dict that contains of table field(s) and its value(s)
        @return: Final dictionary variable contains a table row data.
        """
        final_dict = {}
        for column in self.__table__.columns:

            if type(getattr(self, column.name)) == datetime:
                setattr(self, column.name, getattr(self, column.name).strftime("%Y-%m-%d %H:%M:%S"))

            # final_dict[inflection.underscore(column.name)] = getattr(self, column.name)
            final_dict[column.name] = getattr(self, column.name)
        return final_dict


class SerializeSQLAResult:
    """
    Generating a dict from list of SQL Alchemy result, i.e. eg: <class 'sqlalchemy.util._collections.result'> objects. Finally returns the dict vars in a list.
    """
    # Object variable initialisation
    sqla_util_results = []

    def __init__(self, sqla_util_results):
        """
        Class constructor. Here sqla_util_results is the list contains <class 'sqlalchemy.util._collections.result'> objects
        @param sqla_util_results: list of <class 'sqlalchemy.util._collections.result'> objects/SQL Alchemy result
        """
        self.sqla_util_results = sqla_util_results

    @staticmethod
    def format_columns(dict_row, key, sqla_result, full_date_fields):
        """
        @param dict_row: final dict varible
        @param key: key in the SQLA tuple.
        @param sqla_result: SQLA tuple result.
        @param full_date_fields: List of DateTime fields that needs to be converted to dd-mm-yyyy I:M:S
        (24 Hr with AM/PM) format.
        @return: Formatted dict variable.
        """
        if not key.startswith('_') and key not in ['keys', 'index', 'count', 'metadata', 'query',
                                                   'query_class']:
            # Here the key is a valid column name
            if type(getattr(sqla_result, key)) is datetime or type(getattr(sqla_result, key)) is date:
                # If the column is of datetime object, convert to standard format string value.
                # Eg: 2012-02-23 18:23:33
                # dict_row[key] = getattr(sqla_util_result, key).strftime("%Y-%m-%d %H:%M:%S")
                # Updated
                if full_date_fields:
                    if key in full_date_fields:
                        # For eg: format will be 13-10-2020 04:07:51 PM
                        dict_row[key] = getattr(sqla_result, key).strftime("%d-%m-%Y %I:%M:%S %p")
                else:
                    # Return dd-mm-yyyy format.
                    dict_row[key] = getattr(sqla_result, key).strftime("%d-%m-%Y")
            elif type(getattr(sqla_result, key)) is time:
                # If the column field is time object, convert to standard format string value.
                # Eg: 8:03 AM
                dict_row[key] = getattr(sqla_result, key).strftime("%#I:%M %p")
            elif type(getattr(sqla_result, key)) is Decimal:
                # If the column field is decimal, convert it into a float
                dict_row[key] = float(getattr(sqla_result, key))
            else:
                dict_row[key] = getattr(sqla_result, key)
        return dict_row

    #Added By Athira
    @staticmethod
    def format_columns_date(dict_row, key, sqla_result, full_date_fields):
        """
        @param dict_row: final dict varible
        @param key: key in the SQLA tuple.
        @param sqla_result: SQLA tuple result.
        @param full_date_fields: List of DateTime fields that needs to be converted to dd-mm-yyyy I:M:S
        (24 Hr with AM/PM) format.
        @return: Formatted dict variable.
        """
        if not key.startswith('_') and key not in ['keys', 'index', 'count', 'metadata', 'query',
                                                   'query_class']:
            # Here the key is a valid column name
            if type(getattr(sqla_result, key)) is datetime or type(getattr(sqla_result, key)) is date:
                # If the column is of datetime object, convert to standard format string value.
                # Eg: 2012-02-23 18:23:33
                # dict_row[key] = getattr(sqla_util_result, key).strftime("%Y-%m-%d %H:%M:%S")
                # Updated
                if full_date_fields:
                    if key in full_date_fields:
                        # For eg: format will be Saturday, 13 Feb 2021 20:27:17 PM
                        dict_row[key] = getattr(sqla_result, key).strftime("%a, %d %b %Y %H:%M:%S %p")
                else:
                    # Return dd-mm-yyyy format.
                    dict_row[key] = getattr(sqla_result, key).strftime("%d-%m-%Y")
            elif type(getattr(sqla_result, key)) is time:
                # If the column field is time object, convert to standard format string value.
                # Eg: 8:03 AM
                dict_row[key] = getattr(sqla_result, key).strftime("%#I:%M %p")
            elif type(getattr(sqla_result, key)) is Decimal:
                # If the column field is decimal, convert it into a float
                dict_row[key] = float(getattr(sqla_result, key))
            else:
                dict_row[key] = getattr(sqla_result, key)
        return dict_row

    def serialize_one(self, **kwargs):
        """
        Generating a dict variable from a SQLA result tuple.
        @return: dict variable consisting of SQLA result.
        @rtype:
        """
        dict_row = {}
        full_date_fields = []
        if kwargs.get('full_date_fields') is not None:
            # full_date_fields is a list of DateTime fields that needs to be converted to
            # dd-mm-yyyy h:m:s AM/PM format.
            full_date_fields = kwargs.get('full_date_fields')
        if self.sqla_util_results is not None:
            keys = self.sqla_util_results.keys()
            # Serializing the keys (i.e. columns)
            for key in keys:
                self.format_columns(dict_row, key, self.sqla_util_results, full_date_fields=full_date_fields)

        return dict_row

    def serialize(self, **kwargs):
        """
        Generating the dictionary variable by looping through the list of <class 'sqlalchemy.util._collections.result'> objects
        @return: List of serialized SQLA result.
        """

        full_date_fields = []
        if kwargs.get('full_date_fields') is not None:
            # full_date_fields is a list of DateTime fields that needs to be converted to
            # dd-mm-yyyy h:m:s AM/PM format.
            full_date_fields = kwargs.get('full_date_fields')
        final_list = []
        if self.sqla_util_results is not None:
            for sqla_util_result in self.sqla_util_results:
                dict_row = {}
                keys = sqla_util_result.keys()
                # Serializing the keys (i.e. columns)
                for key in keys:
                    self.format_columns(dict_row, key, sqla_util_result, full_date_fields=full_date_fields)

                # Return variable will be a list type.
                final_list.append(dict_row)
        return final_list

    # Added by Athira
    def serialize_date(self, **kwargs):
        """
        Generating the dictionary variable by looping through the list of <class 'sqlalchemy.util._collections.result'> objects
        @return: List of serialized SQLA result.
        """

        full_date_fields = []
        if kwargs.get('full_date_fields') is not None:
            # full_date_fields is a list of DateTime fields that needs to be converted to
            # dd-mm-yyyy h:m:s AM/PM format.
            full_date_fields = kwargs.get('full_date_fields')
        final_list = []
        if self.sqla_util_results is not None:
            for sqla_util_result in self.sqla_util_results:
                dict_row = {}
                keys = sqla_util_result.keys()
                # Serializing the keys (i.e. columns)
                for key in keys:
                    self.format_columns_date(dict_row, key, sqla_util_result, full_date_fields=full_date_fields)

                # Return variable will be a list type.
                final_list.append(dict_row)
        return final_list


class CallSP:
    """
    This generic class helps to execute stored procedures and populates the result.
    """
    # Class variable for storing the given stored procedure query string.
    query = ''
    # Class variable for storing the result of the stored procedure execution.
    result = None

    def __init__(self, query):
        """
        Setting up the class variable query.
        @param query: query string need to be executed.
        """
        self.query = query

    def execute(self):
        """
        Executing the stored procedure from the given string.
        @return: current CallSP instance.
        """
        try:
            # Executing the stored procedure.
            self.result = db.session.execute(self.query)
        except Exception as e:
            error_logger().error(e)
        return self

    def fetchone(self):
        """
        This function acts as a dummy for the SQLA fetchone() function.
        @return: result of the stored procedure execution.
        """
        if self.result is not None:
            # Calling the SQLA fetchone() function only if the execution of the SP had performed successfully.
            item = self.result.fetchone()
            return_dict = {}
            if item is not None:
                keys = item.keys()
                values = item
                for key, value in zip(keys, values):
                    if type(value) is date:
                        value = value.strftime("%Y-%m-%d")
                    return_dict[key] = value
            # Successfully populated the result. Returning the result as a dict variable.
            return return_dict

        return None

    def fetchall(self):
        """
        This function acts as a dummy for the SQLA fetchall() function.
        @return: result of the stored procedure execution.
        """
        if self.result is not None:
            # Calling the SQLA fetchall() function only if the execution of the SP had performed successfully.
            all_items = self.result.fetchall()
            return_list = []
            if all_items is not None:
                for item in all_items:
                    keys = item.keys()
                    values = item.values()
                    dict_row = {}
                    for key, value in zip(keys, values):
                        # Type conversion for serialization.
                        if type(value) is Decimal:
                            dict_row[key] = float(value)
                        elif type(value) is datetime:
                            dict_row[key] = value.strftime("%d-%m-%Y")
                        elif type(value) is time:
                            dict_row[key] = value.strftime("%#I:%M %p")
                        elif type(value) is date:
                            dict_row[key] = value.strftime("%Y-%m-%d")
                        else:
                            dict_row[key] = value
                    return_list.append(dict_row)
            # Successfully populated the result. Returning the result as a list variable.
            return return_list
        return None

    def fetchall_by_date(self):
        """
        This function acts as a dummy for the SQLA fetchall() function.
        @return: result of the stored procedure execution.
        """
        if self.result is not None:
            # Calling the SQLA fetchall() function only if the execution of the SP had performed successfully.
            all_items = self.result.fetchall()
            return_list = []
            if all_items is not None:
                for item in all_items:
                    keys = item.keys()
                    values = item.values()
                    dict_row = {}
                    for key, value in zip(keys, values):
                        # Type conversion for serialization.
                        if type(value) is Decimal:
                            dict_row[key] = float(value)
                        elif type(value) is datetime:
                            dict_row[key] = value.strftime("%d-%m-%Y %I:%M:%S %p")
                        elif type(value) is time:
                            dict_row[key] = value.strftime("%#I:%M %p")
                        elif type(value) is date:
                            dict_row[key] = value.strftime("%Y-%m-%d")
                        else:
                            dict_row[key] = value
                    return_list.append(dict_row)
            # Successfully populated the result. Returning the result as a list variable.
            return return_list
        return None

class TravelDistanceCalculator:
    """
    A class for calculating the total travelled distance between two GPS co-ordinates.
    Here, the travelled_distance is calculated by adding two adjacent GPS co-ordinates until all
    the positions covered.
    """

    def __init__(self):
        """
        Constructor of the class.
        """
        # Variable initialization.
        self.travelled_distance = 0.0
        self.lat_1 = False
        self.long_1 = False
        self.lat_2 = False
        self.long_2 = False

    def clear(self):
        """
        Function for clearing up the GPS co-ordinate variables.
        @return:
        """
        self.lat_1 = False
        self.long_1 = False
        self.lat_2 = False
        self.long_2 = False

    def loop(self, travel_logs):
        """
        Main function of the class to receive the travel_log variable from the controller.
        @param travel_logs: List of TravelLog model objects.
        @return: current object.
        """
        for log in travel_logs:
            self.calc(log)
        return self

    def calc(self, log):
        """
        Calculating the total travelled distance.
        @param log: The current log that being calculated.
        @return:
        """
        if not self.lat_1 and not self.long_1:
            # In the initial loop, this block will be executed.
            self.lat_1 = log.Lat
            self.long_1 = log.Long

        elif not self.lat_2 and not self.long_2:
            # In the second loop, this block will be executed.
            self.lat_2 = log.Lat
            self.long_2 = log.Long

            # Both lat_1,long_1 and lat_2,long_2 are found.
            # Now the distance between the two can be calculated.
            distance = calculate_distance(self.lat_1, self.long_1, self.lat_2, self.long_2)

            # Adding the current distance between two GPS co-ordinates to the
            # final travelled_distance variable.
            self.travelled_distance += distance

        else:
            # Here, two co-ordinates are already present. So clear up the values.
            self.clear()
            # Repeat the process again. So calling the calc function (recursion).
            self.calc(log)

    def distance(self):
        """
        Function to return the travelled_distance variable.
        @return: travelled_distance
        """
        return round(self.travelled_distance, 2)


class GenerateReport:
    """
    A generic class for generating the excel reports from list of dict variables.
    """

    def __init__(self, list_var, report_name, is_log=False, log_data=None):
        """
        Class constructor. Setting up the list_var
        @param list_var:
        """
        self.report_name = report_name
        self.list_var = list_var
        self.generated_file = None
        self.is_log = is_log
        self.log_data = log_data
        self.root_dir = os.path.dirname(current_app.instance_path)

    def as_text(self, value):
        """
        If the value is None, return "" string.
        @param value:
        @return:
        """
        if value is None:
            return ""
        return str(value)

    def generate(self):
        """
        Generating the excel report based on the list variable.
        @return: this object
        """
        # Creating a Workbook object.
        workbook = Workbook()
        # Setting the header style.
        header = NamedStyle(name="header")
        header.font = Font(bold=True)
        header.border = Border(bottom=Side(border_style="thin"))
        header.alignment = Alignment(horizontal="justify", vertical="justify")

        if self.is_log is False:
            sheet = workbook.active
            # Dict keys will be the sheet's header.
            dict_keys = self.list_var[0].keys()
            keys = []
            for key in dict_keys:
                keys.append(key)
            sheet.append(keys)
            # Selecting the first row as the header row.
            header_row = sheet[1]
            for cell in header_row:
                # Adding this 'header' style to every cell in the header row.
                cell.style = header

            for data in self.list_var:
                row = list([*data.values()])
                # If the list contain any sublist, generate a comma separated string -
                # by concatenating the elements from the sublist.
                for item in row:
                    # Getting the index of the item in the list row.
                    index = row.index(item)
                    if type(item) == list:
                        # Generating a comma separated string from the sublist.
                        item = ','.join(map(str, item))
                        # Replacing the list with the comma separated string.
                        row[index] = item
                    if type(item) == dict:
                        # Converting the dict variable into a string.
                        item = json.dumps(item).replace("{", "").replace("}", "").replace('"', '')
                        # Replacing the dict with the string.
                        row[index] = item

                sheet.append(row)

            # Looping through the cells and setting up the cell width
            # based on length of the text (number of characters) in it.
            for column_cells in sheet.columns:
                length = max(len(self.as_text(cell.value)) for cell in column_cells)
                # Adding an extra width to the original calculation.
                # (To get some space in the cell - for better readability)
                length += 3
                sheet.column_dimensions[column_cells[0].column_letter].width = length

            # Setting up the cell style.
            cell_style = NamedStyle(name="cell_style")
            cell_style.alignment = Alignment(horizontal="justify", vertical="justify")
            # Skipping the header row and apply this style to rest of the columns.
            for row in sheet.iter_rows(min_row=2, max_col=None, max_row=None, values_only=False):
                for cell in row:
                    cell.style = cell_style

        else:
            sheet2 = None
            for rank_list in self.log_data:
                key_name = next(iter(rank_list))

                if len(rank_list[key_name]) != 0:
                    sheet2 = workbook.create_sheet(key_name)
                    # Dict keys will be the sheet's header.
                    dict_keys = rank_list[key_name][0].keys()
                    keys = []
                    for key in dict_keys:
                        keys.append(key)
                    sheet2.append(keys)
                    # Selecting the first row as the header row.
                    header_row = sheet2[1]
                    for cell in header_row:
                        # Adding this 'header' style to every cell in the header row.
                        cell.style = header

                    for data in rank_list[key_name]:
                        row = list([*data.values()])
                        # If the list contain any sublist, generate a comma separated string -
                        # by concatenating the elements from the sublist.
                        for item in row:
                            # Getting the index of the item in the list row.
                            index = row.index(item)
                            if type(item) == list:
                                # Generating a comma separated string from the sublist.
                                item = ','.join(map(str, item))
                                # Replacing the list with the comma separated string.
                                row[index] = item
                            if type(item) == dict:
                                # Converting the dict variable into a string.
                                item = json.dumps(item).replace("{", "").replace("}", "").replace('"', '')
                                # Replacing the dict with the string.
                                row[index] = item

                        sheet2.append(row)

                # Looping through the cells and setting up the cell width
                # based on length of the text (number of characters) in it.
                for column_cells in sheet2.columns:
                    length = max(len(self.as_text(cell.value)) for cell in column_cells)
                    # Adding an extra width to the original calculation.
                    # (To get some space in the cell - for better readability)
                    length += 3
                    sheet2.column_dimensions[column_cells[0].column_letter].width = length

            # Setting up the cell style.
            cell_style = NamedStyle(name="cell_style")
            cell_style.alignment = Alignment(horizontal="justify", vertical="justify")
            # Skipping the header row and apply this style to rest of the columns.
            for row in sheet2.iter_rows(min_row=2, max_col=None, max_row=None, values_only=False):
                for cell in row:
                    cell.style = cell_style

            # Getting the default sheet
            removable_sheet = workbook.get_sheet_by_name('Sheet')
            # Removing the default sheet
            workbook.remove_sheet(removable_sheet)

        # Saving the report.
        file_name = f'{self.report_name}_{datetime.now().strftime("%d-%m-%Y_%I-%M-%S_%p")}'
        workbook.save(filename=f"{self.root_dir}/reports/{file_name}.xlsx")

        target_file = f'{self.root_dir}/reports/{file_name}.xlsx'
        file_exists = os.path.exists(target_file)
        # If the report generated successfully, then set generated_file variable.
        if file_exists:
            self.generated_file = file_name

        return self

    def get(self):
        """
        Function to return the generated excel report.
        @return:
        """
        if self.generated_file is not None:
            return self.generated_file
